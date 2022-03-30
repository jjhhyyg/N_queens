import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill

gray_fill = PatternFill("solid", start_color='C0C0C0')


def clear_wb(wb):
    """
    清空xlsx文件中的内容
    :return: None
    """

    while len(wb.sheetnames) > 1:
        del wb[wb.sheetnames[0]]
    ws = wb.active
    ws.title = 'HelloSheet'


class Checkerboard:
    def __init__(self, wb_name):
        self.wb_name = wb_name

    def play(self, data):
        """
        在xlsx文件中输出N皇后结果
        :param data: 二维列表，每个元素代表一个结果
        :return: None
        """

        # 工作簿对象
        wb = openpyxl.Workbook()
        clear_wb(wb)

        # 结果数
        sheet_num = len(data)
        # 皇后个数（棋盘大小）
        N = len(data[0])

        # 输出num张sheet
        for num in range(sheet_num):
            # 创建sheet并控制范围
            sheet_name = 'Sheet' + str(num + 1)
            ws = wb.create_sheet(sheet_name)
            cells = ws.iter_rows(min_row=1, max_row=N, min_col=1, max_col=N)

            # 表示第几行
            cnt = 0
            for row in cells:
                row[data[num][cnt]].value = '\u2655'
                for cell in row:
                    i, j = cell.row, cell.column
                    coordinate = str(j) + get_column_letter(i)
                    if (i + j) % 2 == 0:
                        cell.fill = gray_fill
                cnt = cnt + 1

        wb.save(self.wb_name)
